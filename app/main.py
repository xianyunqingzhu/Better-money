"""Better-money 本地服务入口：记账、看板、设置 API。

启动：python -m uvicorn app.main:app --host 127.0.0.1 --port 8642
或双击 启动.bat。
"""
import io
import math
import secrets as _secrets
import uuid
from contextlib import asynccontextmanager
from datetime import date, timedelta
from pathlib import Path

from fastapi import FastAPI, File, Form, Header, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from starlette.background import BackgroundTask

from app import ai, backup, db, importers, ledger, recovery, summarizer, uploads
from app import summaries as summary_service
from app.config import load_config, load_raw_config, save_config
from app.data_api import router as data_router
from app.goals import allocate_savings
from app.paths import get_paths, resource_root
from app.version import APP_ID, APP_VERSION, HEALTH_PROTOCOL


@asynccontextmanager
async def lifespan(_: FastAPI):
    get_paths().ensure_directories()
    recovery.recover_interrupted_installs()
    db.init_db()
    conn = db.get_conn()
    try:
        ledger.ensure_finance_config(conn, load_raw_config(), save_config)
    finally:
        conn.close()
    backup.ensure_daily_backup()
    yield


app = FastAPI(title="Better-money", lifespan=lifespan)
app.include_router(data_router)


# ---------- 基础 ----------

@app.get("/api/health")
def health():
    cfg = load_config()
    return {
        "ok": True,
        "app_id": APP_ID,
        "version": APP_VERSION,
        "protocol": HEALTH_PROTOCOL,
        "ai_configured": bool(cfg.get("api_key")),
    }


@app.get("/api/runtime")
def runtime():
    token = getattr(app.state, "session_token", "") or ""
    return JSONResponse(
        {"control_available": bool(token), "session_token": token},
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/control/shutdown")
def control_shutdown(
    x_better_money_token: str | None = Header(default=None),
):
    expected = getattr(app.state, "session_token", "") or ""
    request_shutdown = getattr(app.state, "request_shutdown", None)
    if not expected or request_shutdown is None:
        return JSONResponse(status_code=409, content={
            "error": "shutdown_unavailable",
            "message": "开发者模式由终端管理，无法从这里关闭服务"})
    if not x_better_money_token or not _secrets.compare_digest(
        x_better_money_token, expected
    ):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    # Shut down only after the confirmation response is delivered.
    return JSONResponse(
        {"ok": True}, background=BackgroundTask(request_shutdown))


# ---------- 交易 ----------

class TxIn(BaseModel):
    date: str
    amount: float = Field(gt=0)
    type: str = "支出"
    category: str = "其他"
    merchant: str = ""
    note: str = ""
    source: str = "手动"


def _apply_auto_save(conn, income_item: dict) -> list[dict]:
    """按配置比例规划收入，并返回跨目标的分配明细。"""
    cfg = load_config()
    ratio = float(cfg.get("auto_save_ratio") or 0)
    if ratio <= 0:
        return []
    amount = round(float(income_item["amount"]) * ratio, 2)
    allocations = allocate_savings(conn, amount)
    return [
        {
            "goal_id": allocation.goal_id,
            "goal_name": allocation.goal_name,
            "amount": allocation.amount,
        }
        for allocation in allocations
    ]


@app.post("/api/transactions")
def add_transaction(tx: TxIn):
    conn = db.get_conn()
    cur = conn.execute(
        "INSERT INTO transactions"
        "(date, amount, type, category, merchant, note, source, created_at, updated_at)"
        " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (tx.date, tx.amount, tx.type, tx.category, tx.merchant, tx.note,
         tx.source, db.now_str(), db.now_str()),
    )
    savings_allocations = []
    if tx.type == "收入":
        savings_allocations = _apply_auto_save(conn, {"amount": tx.amount})
    conn.commit()
    tx_id = cur.lastrowid
    conn.close()
    _mark_summaries_expired(tx.date)
    return {
        "ok": True,
        "id": tx_id,
        "savings_allocations": savings_allocations,
    }


@app.get("/api/transactions")
def list_transactions(limit: int = 300):
    conn = db.get_conn()
    rows = conn.execute(
        "SELECT * FROM transactions ORDER BY date DESC, id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.delete("/api/transactions/{tx_id}")
def delete_transaction(tx_id: int):
    conn = db.get_conn()
    row = conn.execute("SELECT date FROM transactions WHERE id = ?", (tx_id,)).fetchone()
    conn.execute("DELETE FROM transactions WHERE id = ?", (tx_id,))
    conn.commit()
    conn.close()
    if row:
        _mark_summaries_expired(row["date"])
    return {"ok": True}


class TxPatch(BaseModel):
    date: str = ""
    amount: float = 0
    type: str = ""
    category: str = ""
    merchant: str | None = None
    note: str | None = None
    estimated: int = -1


@app.patch("/api/transactions/{tx_id}")
def patch_transaction(tx_id: int, p: TxPatch):
    """编辑单笔交易（历史明细页用）。"""
    conn = db.get_conn()
    row = conn.execute("SELECT date FROM transactions WHERE id = ?", (tx_id,)).fetchone()
    if not row:
        conn.close()
        return JSONResponse(status_code=404, content={"error": "not_found", "message": "记录不存在"})
    fields, args = [], []
    if p.date:
        fields.append("date = ?")
        args.append(p.date)
    if p.amount > 0:
        fields.append("amount = ?")
        args.append(p.amount)
    if p.type in ("支出", "收入", "退款", "取现", "转账", "还款"):
        fields.append("type = ?")
        args.append(p.type)
    if p.category:
        fields.append("category = ?")
        args.append(p.category)
    if p.merchant is not None:
        fields.append("merchant = ?")
        args.append(p.merchant)
    if p.note is not None:
        fields.append("note = ?")
        args.append(p.note)
    if p.estimated in (0, 1):
        fields.append("estimated = ?")
        args.append(p.estimated)
    if fields:
        fields.append("updated_at = ?")
        args.append(db.now_str())
        conn.execute(f"UPDATE transactions SET {', '.join(fields)} WHERE id = ?", (*args, tx_id))
    conn.commit()
    conn.close()
    _mark_summaries_expired(p.date or row["date"])
    return {"ok": True}


# ---------- 待处理队列（AI 解析失败保留的内容） ----------

@app.get("/api/pending")
def list_pending():
    conn = db.get_conn()
    rows = conn.execute("SELECT * FROM pending_items ORDER BY id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.delete("/api/pending/{pid}")
def delete_pending(pid: int):
    conn = db.get_conn()
    conn.execute("DELETE FROM pending_items WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------- 数据导出（M7） ----------

@app.get("/api/export/transactions.csv")
def export_transactions_csv():
    conn = db.get_conn()
    rows = conn.execute(
        "SELECT date, amount, type, category, merchant, note, source, estimated "
        "FROM transactions ORDER BY date, id").fetchall()
    conn.close()
    import csv as _csv
    buf = io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["日期", "金额(元)", "类型", "分类", "商家", "备注", "来源", "是否估算"])
    for r in rows:
        w.writerow([r["date"], r["amount"], r["type"], r["category"],
                    r["merchant"], r["note"], r["source"], "是" if r["estimated"] else "否"])
    data = "\ufeff" + buf.getvalue()  # BOM：Excel 打开中文不乱码
    return Response(
        content=data.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="better-money-transactions.csv"'},
    )


def _mark_summaries_expired(date_str: str) -> None:
    """账目变化后，把覆盖该日期的总结（任意区间）标记为过期。"""
    try:
        date.fromisoformat(date_str)
    except (TypeError, ValueError):
        return
    conn = db.get_conn()
    conn.execute(
        "UPDATE summaries SET expired = 1 "
        "WHERE period_start <= ? AND period_end >= ?",
        (date_str, date_str))
    conn.commit()
    conn.close()


# ---------- AI 文字解析（M2） ----------

class ParseIn(BaseModel):
    text: str
    date: str


@app.post("/api/parse_text")
def parse_text(req: ParseIn):
    """把自然语言记账文字解析成多笔交易并直接入账。
    AI 不可用 → 503（前端亮红色横幅）；原文进待处理队列，不丢数据。
    """
    text = req.text.strip()
    if not text:
        return JSONResponse(status_code=400, content={"error": "empty", "message": "内容为空"})

    try:
        result = ai.parse_text(text, req.date)
    except ai.AIUnavailableError as e:
        _store_pending(text)
        return JSONResponse(status_code=503,
                            content={"error": "ai_unavailable", "message": str(e)})
    except Exception as e:
        _store_pending(text)
        return JSONResponse(status_code=500,
                            content={"error": "parse_failed", "message": str(e)})

    items = result.get("items") or []
    questions = result.get("questions") or []
    if not items:
        _store_pending(text)
        return {"ok": True, "saved": 0, "skipped": [], "items": [],
                "questions": questions, "message": "没有识别出可入账的条目"}

    saved, skipped = _save_items(items)
    for it in saved:
        _mark_summaries_expired(it["date"])
    return {"ok": True, "saved": len(saved), "skipped": skipped,
            "items": saved, "questions": questions}


# ---------- 图片识别（M3：小票/截图，确认后再入账） ----------


@app.post("/api/upload_images")
async def upload_images(
    files: list[UploadFile] = File(...),
    note: str = Form(""),
    date: str = Form(""),
):
    """保存图片 → 视觉模型逐张识别 → 返回待确认条目（不直接入账）。"""
    if not files:
        return JSONResponse(status_code=400, content={"error": "empty", "message": "没有文件"})
    if len(files) > uploads.MAX_IMAGES_PER_REQUEST:
        return JSONResponse(status_code=400, content={
            "error": "too_many_files",
            "message": f"一次最多上传 {uploads.MAX_IMAGES_PER_REQUEST} 张图片"})
    day_dir = get_paths().images_dir / (date or "misc")
    day_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    for f in files:
        pending_tmp = None
        try:
            data = await uploads.read_limited(f, uploads.IMAGE_MAX_BYTES)
            ext = uploads.validate_image(data, f.filename or "")
            dest = day_dir / f"{uuid.uuid4().hex}.{ext}"
            pending_tmp = dest.with_name(dest.name + ".uploading")
            pending_tmp.write_bytes(data)
            pending_tmp.replace(dest)
            paths.append(str(dest))
            pending_tmp = None
        except uploads.UploadError as e:
            status = 413 if e.code == "file_too_large" else 400
            return JSONResponse(status_code=status, content={
                "error": e.code, "message": str(e)})
        finally:
            if pending_tmp is not None:
                pending_tmp.unlink(missing_ok=True)

    items, questions, failed, first_err = [], [], 0, ""
    for p in paths:
        try:
            r = ai.parse_image(p, note, date)
            items.extend(r.get("items") or [])
            questions.extend(r.get("questions") or [])
        except ai.AIUnavailableError as e:
            failed += 1
            first_err = str(e)
            _store_pending("", p)
        except Exception as e:
            failed += 1
            first_err = str(e)
            _store_pending("", p)

    if failed == len(paths):
        return JSONResponse(status_code=503, content={
            "error": "ai_unavailable",
            "message": f"图片识别不可用：{first_err}（图片已保存，恢复后可重试）",
        })
    return {"ok": True, "items": items, "questions": questions,
            "images": paths, "failed": failed}


class ConfirmIn(BaseModel):
    items: list[dict]
    source: str = "确认面板"


@app.post("/api/confirm_items")
def confirm_items(req: ConfirmIn):
    """确认面板提交：用户可改过的条目批量入账（含单品明细、去重）。"""
    if not req.items:
        return {"ok": True, "saved": 0, "skipped": []}
    normalized = ai._normalize({"items": req.items}, date.today().isoformat())
    saved, skipped = _save_items(normalized["items"], req.source)
    for it in saved:
        _mark_summaries_expired(it["date"])
    return {"ok": True, "saved": len(saved), "skipped": skipped}


# ---------- 账单 CSV/Excel 导入（M3） ----------

@app.post("/api/import_csv")
async def import_csv(file: UploadFile = File(...)):
    """微信/支付宝账单 CSV(.csv/.xlsx/.xlsm) → 解析为待确认条目（不直接入账）。"""
    try:
        raw = await uploads.read_limited(file, uploads.STATEMENT_MAX_BYTES)
        kind = uploads.validate_statement(raw, file.filename or "")
    except uploads.UploadError as e:
        status = 413 if e.code == "file_too_large" else 400
        return JSONResponse(status_code=status, content={
            "error": e.code, "message": str(e)})

    try:
        if kind in ("xlsx", "xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows = [[("" if c is None else str(c)) for c in row]
                    for row in ws.iter_rows(values_only=True)]
        else:
            text = None
            for enc in ("utf-8-sig", "gbk", "utf-8"):
                try:
                    text = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                return JSONResponse(status_code=400, content={
                    "error": "encoding", "message": "无法识别文件编码，请导出为 CSV 后重试。"})
            import csv as _csv
            rows = list(_csv.reader(io.StringIO(text.lstrip("\ufeff"))))
        items, skipped = importers.parse_rows(rows)
    except Exception as e:
        return JSONResponse(status_code=400, content={
            "error": "format", "message": f"账单文件无法解析：{e}"})
    if not items:
        return {"ok": True, "items": [], "skipped_rows": skipped,
                "message": "没有解析出可入账的交易行"}
    return {"ok": True, "items": items, "skipped_rows": skipped}


def _store_pending(text: str, image_path: str = "") -> None:
    conn = db.get_conn()
    conn.execute("INSERT INTO pending_items(raw_text, image_path, created_at) VALUES (?, ?, ?)",
                 (text, image_path, db.now_str()))
    conn.commit()
    conn.close()


def _save_items(items: list[dict], source: str = "文字"):
    """入账前查重（日期+金额+商家一致视为可能重复，跳过）；含单品明细落库。"""
    conn = db.get_conn()
    saved, skipped = [], []
    for it in items:
        dup = conn.execute(
            "SELECT id FROM transactions WHERE date = ? AND amount = ? AND merchant = ?",
            (it["date"], it["amount"], it["merchant"]),
        ).fetchone()
        if dup:
            skipped.append({**it, "reason": "可能重复"})
            continue
        cur = conn.execute(
            "INSERT INTO transactions"
            "(date, amount, type, category, merchant, note, source, estimated, created_at, updated_at)"
            " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (it["date"], it["amount"], it["type"], it["category"], it["merchant"],
             it["note"], source, it["estimated"], db.now_str(), db.now_str()),
        )
        row = dict(it)
        row["id"] = cur.lastrowid
        if it["type"] == "收入":
            allocations = _apply_auto_save(conn, it)
            row["auto_saved"] = round(
                sum(allocation["amount"] for allocation in allocations), 2
            )
            row["savings_allocations"] = allocations
        for li in it.get("line_items") or []:
            conn.execute(
                "INSERT INTO line_items(transaction_id, name, qty, price) VALUES (?, ?, ?, ?)",
                (row["id"], li.get("name", ""), float(li.get("qty") or 1),
                 float(li.get("price") or 0)),
            )
        saved.append(row)
    conn.commit()
    conn.close()
    return saved, skipped


# ---------- 看板汇总 ----------

def _month_range(month: str):
    """'YYYY-MM' → (first_day, last_day) ISO 字符串；空串 → 当前月。"""
    if month:
        y, m = map(int, month.split("-"))
        first = date(y, m, 1)
    else:
        first = date.today().replace(day=1)
    nxt = (date(first.year + 1, 1, 1) if first.month == 12
           else date(first.year, first.month + 1, 1))
    return first.isoformat(), (nxt - timedelta(days=1)).isoformat()


@app.get("/api/stats")
def stats(month: str = ""):
    """图表数据：所选月的分类占比、近30天趋势、近8周对比。"""
    start, end = _month_range(month)
    cfg = load_config()
    conn = db.get_conn()

    def rows(sql: str, *args):
        return conn.execute(sql, args).fetchall()

    snap = ledger.monthly_snapshot(conn, cfg, f"{start[:7]}")

    # 分类占比（支出 − 退款，按分类）
    cat_rows = rows(
        "SELECT category, SUM(CASE WHEN type='支出' THEN amount ELSE -amount END) AS total "
        "FROM transactions WHERE type IN ('支出','退款') AND date BETWEEN ? AND ? "
        "GROUP BY category HAVING total > 0 ORDER BY total DESC",
        start, end,
    )
    # 月收支
    month_expense = conn.execute(
        "SELECT SUM(CASE WHEN type='支出' THEN amount ELSE -amount END) FROM transactions "
        "WHERE type IN ('支出','退款') AND date BETWEEN ? AND ?", (start, end)
    ).fetchone()[0] or 0
    month_income = conn.execute(
        "SELECT SUM(amount) FROM transactions WHERE type='收入' AND date BETWEEN ? AND ?",
        (start, end),
    ).fetchone()[0] or 0

    # 近30天趋势：结束于今天（当前月）或月末（历史月）
    end_d = date.fromisoformat(end)
    today = date.today()
    trend_end = min(today, end_d)
    trend_start = trend_end - timedelta(days=29)
    daily_rows = rows(
        "SELECT date, SUM(CASE WHEN type='支出' THEN amount ELSE -amount END) AS total "
        "FROM transactions WHERE type IN ('支出','退款') AND date BETWEEN ? AND ? "
        "GROUP BY date",
        trend_start.isoformat(), trend_end.isoformat(),
    )
    daily_map = {r["date"]: round(r["total"], 2) for r in daily_rows}
    daily = []
    d = trend_start
    while d <= trend_end:
        daily.append({"date": d.isoformat(), "value": daily_map.get(d.isoformat(), 0)})
        d += timedelta(days=1)

    # 近8周对比（周一起算）
    this_monday = today - timedelta(days=today.weekday())
    eight_weeks_ago = this_monday - timedelta(days=7 * 7)
    week_rows = rows(
        "SELECT date, SUM(CASE WHEN type='支出' THEN amount ELSE -amount END) AS total "
        "FROM transactions WHERE type IN ('支出','退款') AND date >= ? GROUP BY date",
        eight_weeks_ago.isoformat(),
    )
    by_date = {r["date"]: r["total"] for r in week_rows}
    weekly = []
    for i in range(8):
        ws = eight_weeks_ago + timedelta(days=7 * i)
        we = ws + timedelta(days=6)
        total = 0.0
        dd = ws
        while dd <= we:
            total += by_date.get(dd.isoformat(), 0)
            dd += timedelta(days=1)
        label = "本周" if ws == this_monday else f"{ws.month}/{ws.day}周"
        weekly.append({"label": label, "value": round(total, 2)})

    conn.close()
    return {
        "month": f"{start[:7]}",
        "month_expense": round(month_expense, 2),
        "month_income": round(month_income, 2),
        "opening_balance": round(snap.opening_balance, 2),
        "closing_balance": round(snap.closing_balance, 2),
        "planned_amount": round(snap.planned_amount, 2),
        "unplanned_balance": round(snap.unplanned_balance, 2),
        "category": [{"name": r["category"], "value": round(r["total"], 2)}
                     for r in cat_rows],
        "daily": daily,
        "weekly": weekly,
    }


@app.get("/api/months")
def months():
    """有记录的月份列表（月份切换器用）。"""
    conn = db.get_conn()
    rows = conn.execute(
        "SELECT DISTINCT substr(date, 1, 7) AS m FROM transactions ORDER BY m DESC"
    ).fetchall()
    conn.close()
    return [r["m"] for r in rows]


@app.get("/api/goals")
def list_goals():
    """目标列表（按优先级排序）。"""
    conn = db.get_conn()
    rows = conn.execute("SELECT * FROM goals ORDER BY priority, id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


class GoalIn(BaseModel):
    name: str
    price: float = Field(gt=0)
    expected_date: str = ""
    note: str = ""


@app.post("/api/goals")
def add_goal(g: GoalIn):
    """新增目标 → 默认进入冷静期（cooldown_days 天）。"""
    if not g.name.strip():
        return JSONResponse(status_code=400, content={"error": "bad_name", "message": "名称不能为空"})
    cfg = load_config()
    cd = int(cfg.get("cooldown_days") or 7)
    until = (date.today() + timedelta(days=cd)).isoformat()
    conn = db.get_conn()
    maxp = conn.execute("SELECT COALESCE(MAX(priority), -1) + 1 FROM goals").fetchone()[0]
    cur = conn.execute(
        "INSERT INTO goals(name, price, saved, priority, status, cooldown_until, expected_date, note, created_at) "
        "VALUES (?, ?, 0, ?, '冷静期', ?, ?, ?, ?)",
        (g.name.strip(), g.price, maxp, until, g.expected_date, g.note, db.now_str()))
    conn.commit()
    gid = cur.lastrowid
    conn.close()
    return {"ok": True, "id": gid}


@app.patch("/api/goals/{gid}")
def patch_goal(gid: int, patch: dict):
    """编辑目标字段（含已存金额——手动调拨入口）。"""
    conn = db.get_conn()
    if not conn.execute("SELECT id FROM goals WHERE id = ?", (gid,)).fetchone():
        conn.close()
        return JSONResponse(status_code=404, content={"error": "not_found", "message": "目标不存在"})
    fields, args = [], []
    for k in ("name", "price", "saved", "expected_date", "note"):
        if k in patch:
            fields.append(f"{k} = ?")
            args.append(patch[k])
    if fields:
        conn.execute(f"UPDATE goals SET {', '.join(fields)} WHERE id = ?", (*args, gid))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/api/goals/{gid}")
def delete_goal(gid: int):
    conn = db.get_conn()
    goal = conn.execute(
        "SELECT name, saved FROM goals WHERE id = ?", (gid,)
    ).fetchone()
    if not goal:
        conn.close()
        return JSONResponse(
            status_code=404,
            content={"error": "not_found", "message": "目标不存在"},
        )
    conn.execute("DELETE FROM goals WHERE id = ?", (gid,))
    conn.commit()
    conn.close()
    return {"ok": True, "name": goal["name"], "saved": float(goal["saved"] or 0)}


class GoalAction(BaseModel):
    action: str  # pause/resume/abandon/want/pass/achieve_buy/achieve_freeze/up/down


@app.post("/api/goals/{gid}/action")
def goal_action(gid: int, req: GoalAction):
    conn = db.get_conn()
    g = conn.execute("SELECT * FROM goals WHERE id = ?", (gid,)).fetchone()
    if not g:
        conn.close()
        return JSONResponse(status_code=404, content={"error": "not_found", "message": "目标不存在"})
    act = req.action
    today = date.today().isoformat()
    now = db.now_str()

    if act == "pause":
        conn.execute("UPDATE goals SET status = '已暂停' WHERE id = ?", (gid,))
    elif act == "resume":
        conn.execute("UPDATE goals SET status = '进行中' WHERE id = ?", (gid,))
    elif act == "abandon":
        conn.execute("UPDATE goals SET status = '已放弃', saved = 0 WHERE id = ?", (gid,))
    elif act == "want":  # 冷静期结束，还想要
        conn.execute("UPDATE goals SET status = '进行中' WHERE id = ?", (gid,))
    elif act == "pass":  # 冷静期结束，不想要 → 记入「省下的钱」
        conn.execute("UPDATE goals SET status = '已放弃', saved = 0 WHERE id = ?", (gid,))
        conn.execute(
            "INSERT INTO savings_wins(goal_name, amount, date, created_at) VALUES (?, ?, ?, ?)",
            (g["name"], g["price"], today, now))
    elif act == "achieve_buy":  # 已确认 ②A：记一笔支出 + 标记已达成
        conn.execute("UPDATE goals SET status = '已达成', achieved_at = ? WHERE id = ?", (now, gid))
        conn.execute(
            "INSERT INTO transactions(date, amount, type, category, merchant, note, source, estimated, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (today, g["price"], "支出", "购物", g["name"], "达成目标购买", "目标", 0, now, now))
    elif act == "achieve_freeze":  # 钱够了先不买 → 冻结（暂停）
        conn.execute("UPDATE goals SET status = '已暂停' WHERE id = ?", (gid,))
    elif act in ("up", "down"):
        ordered = conn.execute("SELECT id FROM goals ORDER BY priority, id").fetchall()
        ids = [r["id"] for r in ordered]
        i = ids.index(gid)
        j = i - 1 if act == "up" else i + 1
        if 0 <= j < len(ids):
            ids[i], ids[j] = ids[j], ids[i]
        for k, gid2 in enumerate(ids):
            conn.execute("UPDATE goals SET priority = ? WHERE id = ?", (k, gid2))
    else:
        conn.close()
        return JSONResponse(status_code=400, content={
            "error": "bad_action", "message": f"未知操作 {act}"})

    conn.commit()
    conn.close()
    if act == "achieve_buy":
        _mark_summaries_expired(today)
    return {"ok": True}


class TransferIn(BaseModel):
    to_id: int
    amount: float = Field(gt=0)


@app.post("/api/goals/{gid}/transfer")
def goal_transfer(gid: int, req: TransferIn):
    """目标间调拨已存金额。"""
    conn = db.get_conn()
    frm = conn.execute("SELECT id, saved FROM goals WHERE id = ?", (gid,)).fetchone()
    to = conn.execute("SELECT id FROM goals WHERE id = ?", (req.to_id,)).fetchone()
    if not frm or not to:
        conn.close()
        return JSONResponse(status_code=404, content={"error": "not_found", "message": "目标不存在"})
    if frm["saved"] < req.amount:
        conn.close()
        return JSONResponse(status_code=400, content={
            "error": "not_enough", "message": "该目标已存金额不足"})
    conn.execute("UPDATE goals SET saved = saved - ? WHERE id = ?", (req.amount, gid))
    conn.execute("UPDATE goals SET saved = saved + ? WHERE id = ?", (req.amount, req.to_id))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/api/savings_wins")
def savings_wins(month: str = ""):
    """「省下的钱」：冷静期放弃的购买，按月汇总。"""
    m = month or date.today().strftime("%Y-%m")
    conn = db.get_conn()
    rows = conn.execute(
        "SELECT * FROM savings_wins WHERE date LIKE ? ORDER BY date DESC, id DESC",
        (m + "%",)).fetchall()
    conn.close()
    total = round(sum(r["amount"] for r in rows), 2)
    return {"total": total, "count": len(rows), "items": [dict(r) for r in rows]}


# ---------- 对账（M6） ----------

@app.get("/api/reconcile")
def reconcile_status():
    cfg = load_config()
    conn = db.get_conn()
    bal = _ledger_balance(cfg, conn)
    conn.close()
    return {"ledger_balance": round(bal, 2)}


class ReconcileIn(BaseModel):
    actual: float
    note: str = ""


@app.post("/api/reconcile")
def reconcile(req: ReconcileIn):
    cfg = load_config()
    conn = db.get_conn()
    bal = _ledger_balance(cfg, conn)
    diff = round(req.actual - bal, 2)
    conn.execute(
        "INSERT INTO adjustments(date, diff, note, created_at) VALUES (?, ?, ?, ?)",
        (date.today().isoformat(), diff, req.note or "对账校准", db.now_str()))
    conn.commit()
    conn.close()
    _mark_summaries_expired(date.today().isoformat())
    return {"ok": True, "diff": diff}


@app.get("/api/adjustments")
def list_adjustments():
    """对账调整历史：最新在前，含撤销关系（reversed_by_id）。"""
    conn = db.get_conn()
    rows = conn.execute(
        "SELECT a.*, b.id AS reversed_by_id FROM adjustments a "
        "LEFT JOIN adjustments b ON b.reverses_adjustment_id = a.id "
        "ORDER BY a.created_at DESC, a.id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/adjustments/{aid}/reverse")
def reverse_adjustment(aid: int):
    """撤销一笔对账调整：新增镜像负数记录，不影响原记录。"""
    conn = db.get_conn()
    row = conn.execute(
        "SELECT a.*, b.id AS reversed_by_id FROM adjustments a "
        "LEFT JOIN adjustments b ON b.reverses_adjustment_id = a.id "
        "WHERE a.id = ?", (aid,)).fetchone()
    if not row:
        conn.close()
        return JSONResponse(status_code=404, content={
            "error": "not_found", "message": "调整不存在"})
    if row["reversed_by_id"]:
        conn.close()
        return JSONResponse(status_code=409, content={
            "error": "already_reversed", "message": "该调整已撤销"})
    cur = conn.execute(
        "INSERT INTO adjustments(date, diff, note, created_at, reverses_adjustment_id) "
        "VALUES (?, ?, ?, ?, ?)",
        (row["date"], -row["diff"], f"撤销：{row['note']}", db.now_str(), aid))
    conn.commit()
    conn.close()
    _mark_summaries_expired(row["date"])
    return {"ok": True, "reversal_id": cur.lastrowid}


# ---------- 周/月总结（M5） ----------

@app.get("/api/summaries")
def list_summaries():
    conn = db.get_conn()
    rows = conn.execute("SELECT * FROM summaries ORDER BY created_at DESC, id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


class GenSummaryIn(BaseModel):
    period_type: str                      # 周 / 月（写作风格与篇幅）
    period_start: str = ""                # 开始日期 YYYY-MM-DD
    period_end: str = ""                  # 结束日期 YYYY-MM-DD
    anchor: str = ""                      # 兼容旧调用：未给区间时的锚定日期
    overwrite: bool = False               # 相同区间是否覆盖


@app.post("/api/summaries/generate")
def generate_summary(req: GenSummaryIn):
    start_str = req.period_start or ""
    end_str = req.period_end or ""
    if not start_str or not end_str:
        # 兼容旧调用：未显式给区间时，按锚定日期推导整周/整月
        try:
            anchor_date = date.fromisoformat(req.anchor) if req.anchor else date.today()
        except ValueError:
            anchor_date = date.today()
        start_date, end_date = summarizer.period_range(req.period_type, anchor_date)
        start_str, end_str = start_date.isoformat(), end_date.isoformat()
    try:
        rng = summary_service.SummaryRange.parse(
            start_str, end_str, req.period_type)
    except summary_service.SummaryRangeError as e:
        return JSONResponse(status_code=400, content={
            "error": e.code, "message": str(e)})
    conn = db.get_conn()
    existing_id = summary_service.find_existing(
        conn, rng.period_type, rng.start, rng.end)
    conn.close()
    if existing_id is not None and not req.overwrite:
        return JSONResponse(status_code=409, content={
            "error": "summary_exists",
            "message": "这个类型和区间已经有总结",
            "summary_id": existing_id})
    try:
        content, image_path, summary_id = summarizer.generate(
            rng.period_type, rng.start, rng.end)
    except ai.AIUnavailableError as e:
        return JSONResponse(status_code=503, content={
            "error": "ai_unavailable", "message": str(e)})
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "error": "generate_failed", "message": str(e)})
    return {"ok": True, "content": content, "image_path": image_path,
            "summary_id": summary_id, "overwritten": existing_id is not None}


@app.delete("/api/summaries/{sid}")
def delete_summary(sid: int):
    """删除总结：只影响正文与专属配图，不动账目、目标或对账记录。"""
    try:
        result = summary_service.delete_summary(sid)
    except summary_service.SummaryNotFoundError:
        return JSONResponse(status_code=404, content={
            "error": "not_found", "message": "总结不存在"})
    return {"ok": True, "image_cleanup": result.image_cleanup,
            "message": result.message}


@app.get("/api/summary_image/{sid}")
def summary_image(sid: int):
    conn = db.get_conn()
    row = conn.execute("SELECT image_path FROM summaries WHERE id = ?", (sid,)).fetchone()
    conn.close()
    if not row or not row["image_path"] or not Path(row["image_path"]).exists():
        return JSONResponse(status_code=404, content={"error": "not_found"})
    return FileResponse(row["image_path"])


def _ledger_balance(cfg: dict, conn) -> float:
    """账本余额 = 初始 + 收入 + 退款 − 支出 − 取现/转账/还款 + 对账差额。"""
    start = ledger.initial_balance_start(cfg)
    return ledger.calculate_balance(conn, float(cfg.get("initial_balance") or 0), start)


@app.get("/api/summary")
def summary():
    cfg = load_config()
    conn = db.get_conn()
    month = date.today().strftime("%Y-%m")
    snap = ledger.monthly_snapshot(conn, cfg, month)
    month_start = month + "-01"

    def one(sql: str, *args):
        row = conn.execute(sql, args).fetchone()
        return float(row[0] or 0)

    month_expense = (
        one("SELECT SUM(amount) FROM transactions WHERE type = '支出' AND date >= ?", month_start)
        - one("SELECT SUM(amount) FROM transactions WHERE type = '退款' AND date >= ?", month_start)
    )
    month_income = one(
        "SELECT SUM(amount) FROM transactions WHERE type = '收入' AND date >= ?", month_start
    )
    balance = snap.closing_balance
    conn.close()

    budget = float(cfg["monthly_budget"] or 0)
    ratio = (month_expense / budget) if budget > 0 else 0

    today = date.today()
    next_month = (
        date(today.year + 1, 1, 1) if today.month == 12
        else date(today.year, today.month + 1, 1)
    )
    days_left = (next_month - timedelta(days=1)).day - today.day + 1
    spendable = (budget - month_expense) / max(days_left, 1) if budget > 0 else 0.0

    return {
        "balance": round(balance, 2),
        "opening_balance": round(snap.opening_balance, 2),
        "closing_balance": round(snap.closing_balance, 2),
        "planned_amount": round(snap.planned_amount, 2),
        "unplanned_balance": round(snap.unplanned_balance, 2),
        "month_expense": round(month_expense, 2),
        "month_income": round(month_income, 2),
        "monthly_budget": round(budget, 2),
        "today_spendable": round(spendable, 2),
        "days_left": days_left,
        "budget_ratio": round(ratio, 4),
    }


# ---------- 设置 ----------

@app.get("/api/settings")
def get_settings():
    return load_config()


@app.post("/api/settings")
def set_settings(patch: dict):
    cfg = load_config()
    # 初始余额与初始日期走受保护的专用接口，普通设置不允许修改
    for key in ("initial_balance", "initial_balance_date"):
        patch.pop(key, None)
    cfg.update({k: v for k, v in patch.items() if k in cfg})
    save_config(cfg)
    return {"ok": True}


class InitialBalanceIn(BaseModel):
    initial_balance: float
    initial_balance_date: str


class TestAiIn(BaseModel):
    api_base: str
    api_key: str
    model: str


@app.post("/api/settings/test-ai")
def test_ai_connection(req: TestAiIn):
    """用未保存的草稿值测试 AI 连接；从不记录 API Key。"""
    try:
        ai.test_connection(req.api_base, req.api_key, req.model)
    except ai.AIUnavailableError as e:
        return JSONResponse(status_code=503, content={
            "error": "ai_connection_failed", "message": str(e)})
    return {"ok": True}


@app.post("/api/settings/initial-balance")
def set_initial_balance(req: InitialBalanceIn):
    """受保护的初始余额更正：变更前自动做安全备份。"""
    if not math.isfinite(req.initial_balance):
        return JSONResponse(status_code=400, content={
            "error": "bad_amount", "message": "初始余额必须是有效数字"})
    try:
        new_date = date.fromisoformat(req.initial_balance_date).isoformat()
    except ValueError:
        return JSONResponse(status_code=400, content={
            "error": "bad_date", "message": "日期格式不正确，应为 YYYY-MM-DD"})
    cfg = load_config()
    new_balance = round(req.initial_balance, 2)
    changed = (
        float(cfg.get("initial_balance") or 0) != new_balance
        or str(cfg.get("initial_balance_date") or "") != new_date
    )
    if changed and cfg.get("onboarding_completed"):
        backup.create_backup("pre-initial-balance-change")
    cfg["initial_balance"] = new_balance
    cfg["initial_balance_date"] = new_date
    save_config(cfg)
    conn = db.get_conn()
    balance = ledger.calculate_balance(
        conn, new_balance, date.fromisoformat(new_date))
    conn.close()
    return {"ok": True, "balance": round(balance, 2)}


# ---------- 前端 ----------

app.mount("/static", StaticFiles(directory=resource_root() / "static"), name="static")


@app.get("/")
def index():
    return FileResponse(resource_root() / "static" / "index.html")
